from django.shortcuts import render, redirect, get_object_or_404
from .models import Area, MantenimientoArea, TipoMantenimientoArea
from .forms import AreaForm, MantenimientoAreaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales -----------------------------------------------

@login_required
def area(request):
    alert = Area.objects.all()
    areas = Area.objects.filter(nombre__icontains=request.GET.get('search', ''))
    total_areas = len(areas)
    alertas = []
    for area in alert:
        dias_restantes = area.dias_restantes_mantenimiento()
        if dias_restantes <= 7:
            
            alertas.append({
                'area': area,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'areas': areas,
        'total_areas': total_areas,
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_area/area.html', context)




@login_required
def alertas(request):
    alert = Area.objects.filter(nombre__icontains=request.GET.get('search', ''))
    alertas = []
    for area in alert:
        dias_restantes = area.dias_restantes_mantenimiento()
        if dias_restantes <= 7:
            
            alertas.append({
                'area': area,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_area/alertas.html', context)




@login_required
def tabla_mantenimientos(request):
    areas = Area.objects.all()
    tipos_mantenimiento = TipoMantenimientoArea.objects.all()
    for area in areas:
        area.mantenimientos = area.mantenimientoarea_set.all().order_by('-fecha', '-hora')
    context = {
        'areas': areas,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'SGE_area/tablas.html', context)




@login_required
def crear_area(request):
    if request.method == 'GET':
        form = AreaForm()
        context = {
            'form': form
        }
        return render(request, 'SGE_area/nueva.html', context)
    if request.method == 'POST':
        form = AreaForm(request.POST, request.FILES)  # Asegúrate de pasar request.FILES al formulario
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_area/nueva.html', context)
            else:
                # Manejo del archivo de imagen
                if 'image' in request.FILES:
                    form.instance.image = request.FILES['image']
                form.save()
                return redirect('area')
        else:
            context = {
                'form': form
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_area/nueva.html', context)




@login_required    
def detalles(request, id):
    if request.method == 'GET':
        area = get_object_or_404(Area, id=id)
        mantenimientos = area.mantenimientoarea_set.all().order_by('-fecha', '-hora')
        form = AreaForm(instance=area)
        context = {
            'area': area,
            'form': form,
            'id': id,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_area/detalles.html', context)
    
    if request.method == 'POST':
        area = get_object_or_404(Area, id = id)
        form = AreaForm(instance= area)
        form = AreaForm(request.POST, request.FILES, instance= area)

        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                mantenimientos = area.mantenimientoarea_set.all().order_by('-fecha', '-hora')
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'area': area,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            else:
                form.save()
                mantenimientos = area.mantenimientoarea_set.all().order_by('-fecha', '-hora')
                context = {
                    'area': area,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                return render(request, 'SGE_area/detalles.html', context) 
        
        else:
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url)




@login_required
def eliminar(request, id):
    area = get_object_or_404(Area, id = id)
    area.delete()
    return redirect ('area') 

# fin de vistas generales-----------------------------------------------    




@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoArea, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)




@login_required
def mantenimientos_area_preventivo(request, id):
    if request.method == 'GET':
        area = get_object_or_404(Area, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=2) 
        mantenimientos = area.mantenimientoarea_set.filter(tipo=tipo_mantenimiento).order_by('-fecha', '-hora')
        context = {
            'area': area,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_area/manteniminetos_preventivo.html', context)  




@login_required
def mod_mantenimineto_area_preventivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoArea, id=id)
        area =mantenimiento.area
        form_mant = MantenimientoAreaForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'area': area,
        }
        return render(request, 'SGE_area/mod_mantenimineto_preventivo.html', context)

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoArea, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=2) 
        area =mantenimiento.area
        form_mant = MantenimientoAreaForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.area = area
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_area_preventivo', id=area.id)
        else:
            context = {
                'form_mant': form_mant,
                'area': area,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_area/mod_mantenimineto_preventivo.html', context)    
    
    return HttpResponse("Method Not Allowed", status=405)            




@login_required
def nuevo_mantenimineto_area_preventivo(request, id):
    if request.method == 'GET':
        area = get_object_or_404(Area, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=2) 
        form_mant = MantenimientoAreaForm()
        context = {
            'form_mant': form_mant,
            'area': area,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_area/nuevo_mantenimineto_preventivo.html', context)
    
    if request.method == 'POST':
        area = get_object_or_404(Area, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=2) 
        form_mant = MantenimientoAreaForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.area = area
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_area_preventivo', id=area.id)
        else:
            context = {
                'form_mant': form_mant,
                'area': area,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_area/nuevo_mantenimineto_preventivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)



@login_required
def mantenimientos_area_correctivo(request, id):
    if request.method == 'GET':
        area = get_object_or_404(Area, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=1) 
        mantenimientos = area.mantenimientoarea_set.filter(tipo=tipo_mantenimiento).order_by('-fecha', '-hora')
        context = {
            'area': area,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_area/manteniminetos_correctivo.html', context)  




@login_required
def mod_mantenimineto_area_correctivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoArea, id=id)
        area =mantenimiento.area
        form_mant = MantenimientoAreaForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'area': area,
        }
        return render(request, 'SGE_area/mod_mantenimineto_correctivo.html', context)

    if request.method == 'POST':
            mantenimiento = get_object_or_404(MantenimientoArea, id=id)
            tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=1) 
            area =mantenimiento.area
            form_mant = MantenimientoAreaForm(request.POST, request.FILES, instance=mantenimiento)

            if form_mant.is_valid():
                mantenimiento = form_mant.save(commit=False)
                mantenimiento.area = area
                mantenimiento.tipo = tipo_mantenimiento
                mantenimiento.save()
                return redirect('mantenimientos_area_correctivo', id=area.id)
            else:
                context = {
                    'form_mant': form_mant,
                    'area': area,
                }
                messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
                return render(request, 'SGE_area/mod_mantenimineto_correctivo.html', context)    
    
    return HttpResponse("Method Not Allowed", status=405)            




@login_required
def nuevo_mantenimineto_area_correctivo(request, id):
    if request.method == 'GET':
        area = get_object_or_404(Area, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=1) 
        form_mant = MantenimientoAreaForm()
        context = {
            'form_mant': form_mant,
            'area': area,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_area/nuevo_mantenimineto_correctivo.html', context)
    
    if request.method == 'POST':
        area = get_object_or_404(Area, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=1) 
        form_mant = MantenimientoAreaForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.area = area
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_area_correctivo', id=area.id)
        else:
            context = {
                'form_mant': form_mant,
                'area': area,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_area/nuevo_mantenimineto_correctivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def documento_general_mantenimientos_area(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoArea.objects.filter(fecha__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha__month=mes)

    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha', '-hora')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_areas_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_areas_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Area', 'Tipo', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Operador', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.area.nombre,
            mantenimiento.tipo.tipo,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha,
            mantenimiento.hora,
            mantenimiento.operador,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response 




@login_required
def documento_mantenimientos_preventivos_area(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 2

    area = get_object_or_404(Area, pk=id)
    mantenimientos = MantenimientoArea.objects.filter(area=area).order_by('-fecha', '-hora')

    if mes:
        mantenimientos = mantenimientos.filter(fecha__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}_{}.xlsx"'.format(area.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}.xlsx"'.format(area.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha,
            mantenimiento.hora,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_correctivos_area(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 1

    area = get_object_or_404(Area, pk=id)
    mantenimientos = MantenimientoArea.objects.filter(area=area).order_by('-fecha', '-hora')

    if mes:
        mantenimientos = mantenimientos.filter(fecha__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}_{}.xlsx"'.format(area.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}.xlsx"'.format(area.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha,
            mantenimiento.hora,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response