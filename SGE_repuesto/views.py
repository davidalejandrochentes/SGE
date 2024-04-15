from django.shortcuts import render, redirect, get_object_or_404
from .models import Maquina, Parte, Inventario
from .forms import MaquinaRepuestoForm, ParteRepuestoForm, InventarioRepuestoForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Create your views here.
@login_required
def repuesto_maquina(request):
    if request.method == 'GET':
        maquinas = Maquina.objects.filter(nombre__icontains=request.GET.get('search', ''))
        form = MaquinaRepuestoForm()
        context = {
            'maquinas': maquinas,
            'form': form
        }
        return render(request, 'SGE_repuesto/repuesto.html', context)
    if request.method == 'POST':
        form = MaquinaRepuestoForm(request.POST)
        if form.is_valid():
            form.save()
            form = MaquinaRepuestoForm()
            maquinas = Maquina.objects.filter(nombre__icontains=request.GET.get('search', ''))
            context = {
            'maquinas': maquinas,
            'form': form
            }
            return render(request, 'SGE_repuesto/repuesto.html', context)
        else:
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url)   
        

@login_required
def eliminar_repuesto_maquina(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    maquina.delete()
    return redirect('repuesto_maquina')

@login_required
def eliminar_inventario(request, id):
    inventario = get_object_or_404(Inventario, id=id)
    inventario.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)    


def actualizar_inventario(request, id):
    inventario = get_object_or_404(Inventario, id=id)
    if request.method == 'POST':
        form = InventarioRepuestoForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url)
    else:
        form = InventarioRepuestoForm(instance=inventario)
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)

@login_required
def detalles(request, id):
    if request.method == 'GET':
        maquina = get_object_or_404(Maquina, id=id)
        partes_maquina = Parte.objects.filter(maquina=maquina)
        part_form = ParteRepuestoForm()
        inve_form = InventarioRepuestoForm()
        context = {
            'maquina': maquina,
            'partes_maquina': partes_maquina,
            'part_form': part_form,
            'inve_form': inve_form
        }
        return render(request, 'SGE_repuesto/detalles.html', context) 

    if request.method == 'POST':
        maquina = get_object_or_404(Maquina, id=id)
        partes_maquina = Parte.objects.filter(maquina=maquina)
        part_form = ParteRepuestoForm(request.POST)
        inve_form = InventarioRepuestoForm(request.POST)

        action = request.POST.get('action')
        if action == 'delete':
            parte_id = request.POST.get('parte')
            parte_instance = get_object_or_404(Parte, id=parte_id)
            parte_instance.delete()
            
        if part_form.is_valid():
            parte = part_form.save(commit=False)
            parte.maquina = maquina
            parte.save()

        if inve_form.is_valid():
            inve_instance = inve_form.save(commit=False)
            parte_id = request.POST.get('parte')
            parte_instance = get_object_or_404(Parte, id=parte_id)
            inve_instance.parte = parte_instance
            inve_instance.save()    

        maquina = get_object_or_404(Maquina, id=id)
        part_form = ParteRepuestoForm()
        inve_form = InventarioRepuestoForm()
        partes_maquina = Parte.objects.filter(maquina=maquina)
        context = {
            'maquina': maquina,
            'part_form': part_form,
            'partes_maquina': partes_maquina,
            'inve_form': inve_form,
        }
        return render(request, 'SGE_repuesto/detalles.html', context)  



def descargar_excel(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    partes_maquina = Parte.objects.filter(maquina=maquina)

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalles"

    header_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")

    headers = ["Parte", "Tipo", "Rosca", "Largo", "Und", "Cantidad necesaria", "Existencia en stock", "Salida", "Existencia física"]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill

    for parte in partes_maquina:
        inventarios = parte.inventario_set.all()
        if inventarios:
            start_row = ws.max_row + 1
            end_row = start_row + len(inventarios) - 1
            ws.cell(row=start_row, column=1, value=parte.nombre)
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
            for i, inventario in enumerate(inventarios, start=start_row):
                row_data = [inventario.tipo, inventario.rosca, inventario.largo, inventario.und, inventario.cantidad_necesaria, inventario.existencia_stock, inventario.salida, inventario.existencia_fisica()]
                for j, value in enumerate(row_data, start=2):  # Start from column 2 to leave the first column for "Parte"
                    ws.cell(row=i, column=j, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="detalles_maquina_{maquina.nombre}.xlsx"'
    wb.save(response)

    return response



def tabla_general(request):
    inventarios = Inventario.objects.all()
    context = {
        'inventarios': inventarios,
    }
    return render(request, 'SGE_repuesto/tabla.html', context)


def descargar_excel_general(request):
    inventarios = Inventario.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalles"

    header_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")

    headers = ["Maquina", "Parte de la máquina", "Tipo", "Rosca", "Largo", "Und", "Cantidad necesaria", "Existencia en stock", "Salida", "Existencia física"]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill

    current_maquina = None
    current_parte = None
    start_row_maquina = 2  # Start from the second row (after headers) for maquina
    start_row_parte = 2  # Start from the second row (after headers) for parte
    for inventario in inventarios:
        row_data = [inventario.parte.maquina.nombre, inventario.parte.nombre, inventario.tipo, inventario.rosca, inventario.largo, inventario.und, inventario.cantidad_necesaria, inventario.existencia_stock, inventario.salida, inventario.existencia_fisica()]
        
        # Handling Maquina column
        if inventario.parte.maquina != current_maquina:
            if current_maquina is not None:
                end_row_maquina = ws.max_row
                ws.merge_cells(start_row=start_row_maquina, end_row=end_row_maquina, start_column=1, end_column=1)  # Merge cells for the current maquina
            current_maquina = inventario.parte.maquina
            start_row_maquina = ws.max_row + 1
          
        # Handling Parte de la máquina column
        if inventario.parte != current_parte:
            if current_parte is not None:
                end_row_parte = ws.max_row
                ws.merge_cells(start_row=start_row_parte, end_row=end_row_parte, start_column=2, end_column=2)  # Merge cells for the current parte
            current_parte = inventario.parte
            start_row_parte = ws.max_row + 1
            
        ws.append(row_data)

    # Merge cells for the last maquina and parte
    if current_maquina is not None:
        end_row_maquina = ws.max_row
        ws.merge_cells(start_row=start_row_maquina, end_row=end_row_maquina, start_column=1, end_column=1)
    if current_parte is not None:
        end_row_parte = ws.max_row
        ws.merge_cells(start_row=start_row_parte, end_row=end_row_parte, start_column=2, end_column=2)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="detalles_inventario.xlsx"'
    wb.save(response)

    return response
