from django.shortcuts import render, redirect, get_object_or_404
from .models import Herramienta, MantenimientoHerramienta, TipoMantenimientoHerramienta
from .forms import HerramientaForm, MantenimientoHerramientaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, PatternFill

# Create your views here.
@login_required
def herramienta(request):
    alert = Herramienta.objects.all()
    herramientas = Herramienta.objects.filter(nombre__icontains=request.GET.get('search', ''))
    total_herramientas = len(herramientas)
    alertas = []
    for herramienta in alert:
        dias_restantes = herramienta.dias_restantes_mantenimiento()
        if dias_restantes <= 7:
            
            alertas.append({
                'herramienta': herramienta,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'herramientas': herramientas,
        'total_herramientas': total_herramientas,
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_herramienta/herramienta.html', context)    

@login_required
def alertas(request):
    alert = Herramienta.objects.filter(nombre__icontains=request.GET.get('search', ''))
    alertas = []
    for herramienta in alert:
        dias_restantes = herramienta.dias_restantes_mantenimiento()
        if dias_restantes <= 7:
            
            alertas.append({
                'herramienta': herramienta,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_herramienta/alertas.html', context)    

@login_required
def tabla_mantenimientos(request):
    herramientas = Herramienta.objects.all()
    tipos_mantenimiento = TipoMantenimientoHerramienta.objects.all()
    for herramienta in herramientas:
        herramienta.mantenimientos = herramienta.mantenimientoherramienta_set.all().order_by('-fecha', '-hora')
    context = {
        'herramientas': herramientas,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'SGE_herramienta/tablas.html', context)    

''' 
@login_required
def crear_herramienta(request):
    if request.method == 'GET':
        form = HerramientaForm()
        context = {
            'form': form
        }
        return render(request, 'SGE_herramienta/nueva.html', context)
    if request.method == 'POST':
        form = HerramientaForm(request.POST, request.FILES)  # Asegúrate de pasar request.FILES al formulario
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_herramienta/nueva.html', context)
            else:
                # Manejo del archivo de imagen
                if 'image' in request.FILES:
                    form.instance.image = request.FILES['image']

                form.save()
                return redirect('herramienta')
        else:
            context = {
                'form': form
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_herramienta/nueva.html', context)  
'''  
@login_required
def crear_herramienta(request):
    if request.method == 'GET':
        form = HerramientaForm()
        context = {
            'form': form
        }
        return render(request, 'SGE_herramienta/nueva.html', context)
    if request.method == 'POST':
        form = HerramientaForm(request.POST, request.FILES)  # Asegúrate de pasar request.FILES al formulario
        if form.is_valid():
            form.save()
            return redirect('herramienta')
        else:
            context = {
                'form': form
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_herramienta/nueva.html', context)             

@login_required
def eliminar(request, id):
    herramienta = get_object_or_404(Herramienta, id = id)
    herramienta.delete()
    return redirect ('herramienta') 

@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoHerramienta, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)    

@login_required    
def detalles(request, id):
    if request.method == 'GET':
        herramienta = get_object_or_404(Herramienta, id=id)
        mantenimientos = herramienta.mantenimientoherramienta_set.all().order_by('-fecha', '-hora')
        form = HerramientaForm(instance=herramienta)
        form_mant = MantenimientoHerramientaForm()
        tipos_mantenimiento = TipoMantenimientoHerramienta.objects.all()
        context = {
            'herramienta': herramienta,
            'form': form,
            'id': id,
            'form_mant': form_mant,
            'mantenimientos': mantenimientos,
            'tipos_mantenimiento': tipos_mantenimiento,
            }
        return render(request, 'SGE_herramienta/detalles.html', context)
    
    if request.method == 'POST':
        herramienta = get_object_or_404(Herramienta, id = id)
        form_mant = MantenimientoHerramientaForm(request.POST)
        form = HerramientaForm(instance= herramienta)
        form = HerramientaForm(request.POST, request.FILES, instance= herramienta)

        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form_mant = MantenimientoHerramientaForm()
                tipos_mantenimiento = TipoMantenimientoHerramienta.objects.all()
                mantenimientos = herramienta.mantenimientoherramienta_set.all().order_by('-fecha', '-hora')
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'herramienta': herramienta,
                    'form': form,
                    'id': id,
                    'form_mant': form_mant,
                    'mantenimientos': mantenimientos,
                    'tipos_mantenimiento': tipos_mantenimiento,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            else:
                form.save()
                form_mant = MantenimientoHerramientaForm()
                tipos_mantenimiento = TipoMantenimientoHerramienta.objects.all()
                mantenimientos = herramienta.mantenimientoherramienta_set.all().order_by('-fecha', '-hora')
                context = {
                    'herramienta': herramienta,
                    'form': form,
                    'id': id,
                    'form_mant': form_mant,
                    'mantenimientos': mantenimientos,
                    'tipos_mantenimiento': tipos_mantenimiento,
                }
                return render(request, 'SGE_herramienta/detalles.html', context) 
        
        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.herramienta = herramienta
            mantenimiento.save()
            form = HerramientaForm(instance= herramienta)
            tipos_mantenimiento = TipoMantenimientoHerramienta.objects.all()
            mantenimientos = herramienta.mantenimientoherramienta_set.all().order_by('-fecha', '-hora')
            context = {
            'herramienta': herramienta,
            'form': form,
            'id': id,
            'form_mant': form_mant,
            'mantenimientos': mantenimientos,
            'tipos_mantenimiento': tipos_mantenimiento,
            }
            return render(request, 'SGE_herramienta/detalles.html', context)  
        else:
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url)            



def descargar_herramientas(request):
    herramientas = Herramienta.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="herramientas.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Nombre', 'Número de serie', 'Encargado', 'Teléfono encargado', 'Descripción', 'Fecha de adquisición', 'Costo', 'Proveedor', 'Ubicación', 'Estado de la herramienta', 'Fecha último mantenimiento', 'Intervalo mantenimiento']
    
    # Configuración de estilos para la cabecera
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    
    # Inserción de datos
    for row, herramienta in enumerate(herramientas, start=2):
        ws.cell(row=row, column=1, value=herramienta.nombre)
        ws.cell(row=row, column=2, value=herramienta.número_de_serie)
        ws.cell(row=row, column=3, value=herramienta.encargado)
        ws.cell(row=row, column=4, value=herramienta.teléfono_encargado)
        ws.cell(row=row, column=5, value=herramienta.descripción)
        ws.cell(row=row, column=6, value=herramienta.fecha_de_adquisición)
        ws.cell(row=row, column=7, value=herramienta.costo)
        ws.cell(row=row, column=8, value=herramienta.proveedor)
        ws.cell(row=row, column=9, value=herramienta.ubicación)
        ws.cell(row=row, column=10, value=herramienta.estado_de_la_herramienta)
        ws.cell(row=row, column=11, value=herramienta.fecha_ultimo_mantenimiento)
        ws.cell(row=row, column=12, value=herramienta.intervalo_mantenimiento)

    # Ajuste del ancho de las columnas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(response)
    return response
