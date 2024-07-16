from django.shortcuts import render, redirect, get_object_or_404
from .models import Vehiculo, MantenimientoVehiculo, TipoMantenimientoVehiculo, Viaje
from .forms import VehiculoForm, MantenimientoVehiculoCorrectivoForm, MantenimientoVehiculoPreventivoForm, ViajeVehiculoForm, ViajeVehiculoModAdminForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.forms import AuthenticationForm

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales ----------------------------------------------------------------------------

@login_required
def vehiculo(request):
    alert = Vehiculo.objects.all()
    vehiculos = Vehiculo.objects.filter(marca__icontains=request.GET.get('search', ''))
    total_vehiculos = len(vehiculos)
    alertas = []
    for vehiculo in alert:

        # Cambio de filtro de aceite
        km_restantes_filtro_aceite = vehiculo.km_restantes_cambio_de_filtro_aceite()
        if km_restantes_filtro_aceite <= 500:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aceite,
            })

        # Cambio de filtro de aire/combustible
        km_restantes_filtro_aire_combustible = vehiculo.km_restantes_cambio_filtro_aire_combustible()
        if km_restantes_filtro_aire_combustible <= 500:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aire_combustible,
            })

        # Cambio de filtro de caja/corona
        km_restantes_filtro_caja_corona = vehiculo.km_restantes_cambio_filtro_caja_corona()
        if km_restantes_filtro_caja_corona <= 500:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_caja_corona,
            })
    
    alertas_ordenadas = sorted(alertas, key=lambda x: x['km_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'vehiculos': vehiculos,
        'total_vehiculos': total_vehiculos,
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_vehiculo/vehiculo.html', context)




@login_required
def alertas(request):
    search_query = request.GET.get('search', '')
    vehiculos = Vehiculo.objects.filter(marca__icontains=search_query)
    alertas = []

    for vehiculo in vehiculos:
        # Cambio de filtro de aceite
        km_restantes_filtro_aceite = vehiculo.km_restantes_cambio_de_filtro_aceite()
        if km_restantes_filtro_aceite <= 500:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aceite,
                'tipo': TipoMantenimientoVehiculo.objects.get(id=3).tipo
            })
        # Cambio de filtro de aire/combustible
        km_restantes_filtro_aire_combustible = vehiculo.km_restantes_cambio_filtro_aire_combustible()
        if km_restantes_filtro_aire_combustible <= 500:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aire_combustible,
                'tipo': TipoMantenimientoVehiculo.objects.get(id=4).tipo
            })
        # Cambio de filtro de caja/corona
        km_restantes_filtro_caja_corona = vehiculo.km_restantes_cambio_filtro_caja_corona()
        if km_restantes_filtro_caja_corona <= 500:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_caja_corona,
                'tipo': TipoMantenimientoVehiculo.objects.get(id=5).tipo
            })

    alertas_ordenadas = sorted(alertas, key=lambda x: x['km_restantes'])
    total_alertas = len(alertas_ordenadas)
    
    context = {
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_vehiculo/alertas.html', context)




@login_required
def tabla_mantenimientos(request):
    vehiculos = Vehiculo.objects.all()
    tipos_mantenimiento = TipoMantenimientoVehiculo.objects.all()
    for vehiculo in vehiculos:
        vehiculo.mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
    context = {
        'vehiculos': vehiculos,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'SGE_vehiculo/tablas.html', context) 




@login_required
def crear_vehiculo(request):
    if request.method == 'GET':
        form = VehiculoForm()
        context = {
            'form': form
        }
        return render(request, 'SGE_vehiculo/nuevo.html', context)
    if request.method == 'POST':
        form = VehiculoForm(request.POST, request.FILES)  # Asegúrate de pasar request.FILES al formulario
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_vehiculo/nuevo.html', context)
            else:
                if 'image' in request.FILES:
                    form.instance.image = request.FILES['image']
                form.save()
                return redirect('vehiculo')
        else:
            context = {
                'form': form
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_vehiculo/nuevo.html', context)    




@login_required    
def detalles(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
        form = VehiculoForm(instance=vehiculo)
        context = {
            'vehiculo': vehiculo,
            'form': form,
            'id': id,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/detalles.html', context)
    
    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        form = VehiculoForm(instance=vehiculo)
        form = VehiculoForm(request.POST, request.FILES, instance=vehiculo)

        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            else:
                form.save()
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                return render(request, 'SGE_vehiculo/detalles.html', context) 
        
        else:
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url) 




@login_required
def eliminar(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    vehiculo.delete()
    return redirect ('vehiculo')    
# fin de vistas generales----------------------------------------------------------------------------------



#viajes---------------------------------------------------------------------------------------------------
def viaje(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    viajes = vehiculo.viaje_set.all().order_by('-fecha_llegada', '-hora_llegada')
    ultimo_viaje = vehiculo.viaje_set.last()
    context = {
        'vehiculo': vehiculo,
        'viajes': viajes,
        'ultimo_viaje': ultimo_viaje,
    }
    return render(request, 'SGE_vehiculo/viajes.html', context)




def log_in_vehiculo(request):
    if request.method =='GET':
        return render(request, 'SGE_vehiculo/log.html', {'form': AuthenticationForm})
    else:
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            vehiculo = Vehiculo.objects.get(nombre_usuario_chofer=username, contraseña_chofer=password)
            return redirect('viaje', id=vehiculo.id)
        except Vehiculo.DoesNotExist:
            messages.error(request, "El usuario no existe, o la contraseña es incorrecta")
            return render(request, 'SGE_vehiculo/log.html', {'form': AuthenticationForm})




def nuevo_viaje_vehiculo(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        initial_data = {
            'kilometraje_de_salida': vehiculo.km_recorridos
        }
        viaje_form = ViajeVehiculoForm(initial=initial_data)
    
        context = {
            'viaje_form': viaje_form,
            'vehiculo': vehiculo,
        }
        return render(request, 'SGE_vehiculo/nuevo_viaje.html', context)

    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        viaje_form = ViajeVehiculoForm(request.POST, request.FILES)

        if viaje_form.is_valid():
            viaje = viaje_form.save(commit=False)
            viaje.vehiculo = vehiculo
            viaje.kilometraje_de_llegada = int(request.POST.get('kilometraje_de_salida'))
            viaje.save()
            return redirect('viaje', id=vehiculo.id)
        else:
            context = {
                'viaje_form': viaje_form,
                'vehiculo': vehiculo,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/nuevo_viaje.html', context)

    return HttpResponse("Method Not Allowed", status=405)




def mod_viaje_vehiculo_admin(request, id):
    if request.method == 'GET':
        viaje = get_object_or_404(Viaje, id=id)
        vehiculo = viaje.vehiculo
        form_viaje = ViajeVehiculoModAdminForm(instance=viaje)
        context = {
            'form_viaje': form_viaje,
            'vehiculo': vehiculo,
        }
        return render(request, 'SGE_vehiculo/mod_viaje.html', context)

    if request.method == 'POST':
        viaje = get_object_or_404(Viaje, id=id)
        vehiculo = viaje.vehiculo
        form_viaje = ViajeVehiculoModAdminForm(request.POST, request.FILES, instance=viaje)

        if form_viaje.is_valid():
            viaje = form_viaje.save(commit=False)
            viaje.vehiculo = vehiculo
            viaje.save()
            return redirect('viaje', id=vehiculo.id)
        else:
            context = {
                'form_viaje': form_viaje,
                'vehiculo': vehiculo,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/mod_viaje.html', context)

    return HttpResponse("Method Not Allowed", status=405)    




@login_required
def eliminar_viaje(request, id):
    viaje = get_object_or_404(Viaje, id=id)
    viaje.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)
#---------------------------------------------------------------------------------------------------------------



@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)




@login_required
def mantenimientos_vehiculo_preventivo(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=2)
        mantenimientos = vehiculo.mantenimientovehiculo_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/manteniminetos_preventivo.html', context)




@login_required
def mod_mantenimineto_vehiculo_preventivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
        }
        return render(request, 'SGE_vehiculo/mod_mantenimineto_preventivo.html', context)

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=2)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_preventivo', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/mod_mantenimineto_preventivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def nuevo_mantenimineto_vehiculo_preventivo(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=2)
        form_mant = MantenimientoVehiculoPreventivoForm()
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_vehiculo/nuevo_mantenimineto_preventivo.html', context)

    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=2)
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_preventivo', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/nuevo_mantenimineto_preventivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def mantenimientos_vehiculo_correctivo(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=1)
        mantenimientos = vehiculo.mantenimientovehiculo_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/manteniminetos_correctivo.html', context)




@login_required
def mod_mantenimineto_vehiculo_correctivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoCorrectivoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
        }
        return render(request, 'SGE_vehiculo/mod_mantenimineto_correctivo.html', context)
    
    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=1)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoCorrectivoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_correctivo', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/mod_mantenimineto_correctivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def nuevo_mantenimineto_vehiculo_correctivo(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=1)
        form_mant = MantenimientoVehiculoCorrectivoForm()
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_vehiculo/nuevo_mantenimineto_correctivo.html', context)
    
    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=1)
        form_mant = MantenimientoVehiculoCorrectivoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_correctivo', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/nuevo_mantenimineto_correctivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def mantenimientos_vehiculo_cambio_filtro_aceite(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=3)
        mantenimientos = vehiculo.mantenimientovehiculo_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/manteniminetos_cambio_filtro_aceite.html', context)




@login_required
def mod_mantenimineto_vehiculo_cambio_filtro_aceite(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
        }
        return render(request, 'SGE_vehiculo/mod_mantenimineto_cambio_filtro_aceite.html', context)

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=3)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_cambio_filtro_aceite', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/mod_mantenimineto_cambio_filtro_aceite.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def nuevo_mantenimineto_vehiculo_cambio_filtro_aceite(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=3)
        form_mant = MantenimientoVehiculoPreventivoForm()
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_vehiculo/nuevo_mantenimineto_cambio_filtro_aceite.html', context)

    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=3)
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_cambio_filtro_aceite', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/nuevo_mantenimineto_cambio_filtro_aceite.html', context)

    return HttpResponse("Method Not Allowed", status=405)    




@login_required
def mantenimientos_vehiculo_cambio_filtro_aire_combustible(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=4)
        mantenimientos = vehiculo.mantenimientovehiculo_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/manteniminetos_cambio_filtro_aire_combustible.html', context)




@login_required
def mod_mantenimineto_vehiculo_cambio_filtro_aire_combustible(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
        }
        return render(request, 'SGE_vehiculo/mod_mantenimineto_cambio_filtro_aire_combustible.html', context)

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=4)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_cambio_filtro_aire_combustible', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/mod_mantenimineto_cambio_filtro_aire_combustible.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def nuevo_mantenimineto_vehiculo_cambio_filtro_aire_combustible(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=4)
        form_mant = MantenimientoVehiculoPreventivoForm()
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_vehiculo/nuevo_mantenimineto_cambio_filtro_aire_combustible.html', context)

    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=4)
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_cambio_filtro_aire_combustible', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/nuevo_mantenimineto_cambio_filtro_aire_combustible.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def mantenimientos_vehiculo_cambio_filtro_caja_corona(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=5)
        mantenimientos = vehiculo.mantenimientovehiculo_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/manteniminetos_cambio_filtro_caja_corona.html', context)




@login_required
def mod_mantenimineto_vehiculo_cambio_filtro_caja_corona(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
        }
        return render(request, 'SGE_vehiculo/mod_mantenimineto_cambio_filtro_caja_corona.html', context)

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=5)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_cambio_filtro_caja_corona', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/mod_mantenimineto_cambio_filtro_caja_corona.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def nuevo_mantenimineto_vehiculo_cambio_filtro_caja_corona(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=5)
        form_mant = MantenimientoVehiculoPreventivoForm()
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_vehiculo/nuevo_mantenimineto_cambio_filtro_caja_corona.html', context)

    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=5)
        form_mant = MantenimientoVehiculoPreventivoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            mantenimiento.save()
            return redirect('mantenimientos_vehiculo_cambio_filtro_caja_corona', id=vehiculo.id)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/nuevo_mantenimineto_cambio_filtro_caja_corona.html', context)

    return HttpResponse("Method Not Allowed", status=405)  




# descargas -----------------------------------------------------------------------------------




@login_required
def documento_general_mantenimientos_vehiculo(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoVehiculo.objects.filter(fecha_fin__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha__month=mes)

    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha_fin', '-hora_fin')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_vehiculos_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_vehiculos_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Marca', 'Modelo', 'Tipo', 'Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.vehiculo.marca,
            mantenimiento.vehiculo.modelo,
            mantenimiento.tipo.tipo,
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_preventivos_vehiculo(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 2

    vehiculo = get_object_or_404(Vehiculo, pk=id)
    mantenimientos = MantenimientoVehiculo.objects.filter(vehiculo=vehiculo).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_correctivos_vehiculo(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 1

    vehiculo = get_object_or_404(Vehiculo, pk=id)
    mantenimientos = MantenimientoVehiculo.objects.filter(vehiculo=vehiculo).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_cambio_filtro_aceite_vehiculo(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 3

    vehiculo = get_object_or_404(Vehiculo, pk=id)
    mantenimientos = MantenimientoVehiculo.objects.filter(vehiculo=vehiculo).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_cambio_filtro_aceite_de_{}_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_cambio_filtro_aceite_de_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_cambio_filtro_aire_combustible_vehiculo(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 4

    vehiculo = get_object_or_404(Vehiculo, pk=id)
    mantenimientos = MantenimientoVehiculo.objects.filter(vehiculo=vehiculo).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_cambio_filtro_aire_combustible_de_{}_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_cambio_filtro_aire_combustible_de_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_cambio_filtro_caja_corona_vehiculo(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 5

    vehiculo = get_object_or_404(Vehiculo, pk=id)
    mantenimientos = MantenimientoVehiculo.objects.filter(vehiculo=vehiculo).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_cambio_filtro_caja_corona_de_{}_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_cambio_filtro_caja_corona_de_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_viajes_vehiculo(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')

    vehiculo = get_object_or_404(Vehiculo, pk=id)
    viajes = Viaje.objects.filter(vehiculo=vehiculo).order_by('-fecha_llegada', '-hora_llegada')

    if mes:
        viajes = viajes.filter(fecha_llegada__month=mes)
    if anio:
        viajes = viajes.filter(fecha_llegada__year=anio)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="viajes_de_{}_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="viajes_de_{}_{}_{}.xlsx"'.format(vehiculo.marca, vehiculo.modelo, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Origen', 'Destino', 'Fecha de salida', 'Hora de salida', 'Kilometraje de salida', 'Fecha de llegada', 'Hora de llegada', 'Kilometraje de llegada']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for viaje in viajes:
        ws.append([
            viaje.origen,
            viaje.destino,
            viaje.fecha_salida,
            viaje.hora_salida,
            viaje.kilometraje_de_salida,
            viaje.fecha_llegada,
            viaje.hora_llegada,
            viaje.kilometraje_de_llegada,
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response