from django.shortcuts import render, redirect, get_object_or_404
from .models import PC, MantenimientoPC, TipoMantenimientoPC
from .forms import PCForm, MantenimientoPCForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales ----------------------------------------------------------

@login_required
def pc(request):
    alert = PC.objects.all()
    pcs = PC.objects.filter(nombre__icontains=request.GET.get('search', ''))
    total_pcs = len(pcs)
    alertas = []
    for pc in alert:
        dias_restantes = pc.dias_restantes_mantenimiento()
        if dias_restantes <= 7:
            
            alertas.append({
                'pc': pc,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'pcs': pcs,
        'total_pcs': total_pcs,
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_pc/pc.html', context)




@login_required
def alertas(request):
    alert = PC.objects.filter(nombre__icontains=request.GET.get('search', ''))
    alertas = []
    for pc in alert:
        dias_restantes = pc.dias_restantes_mantenimiento()
        if dias_restantes <= 7:
            
            alertas.append({
                'pc': pc,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_pc/alertas.html', context)




@login_required
def tabla_mantenimientos(request):
    pcs = PC.objects.all()
    tipos_mantenimiento = TipoMantenimientoPC.objects.all()
    for pc in pcs:
        pc.mantenimientos = pc.mantenimientopc_set.all().order_by('-fecha', '-hora')
    context = {
        'pcs': pcs,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'SGE_pc/tablas.html', context)




@login_required
def crear_pc(request):
    if request.method == 'GET':
        form = PCForm()
        context = {
            'form': form
        }
        return render(request, 'SGE_pc/nueva.html', context)
    if request.method == 'POST':
        form = PCForm(request.POST, request.FILES)  # Asegúrate de pasar request.FILES al formulario
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_pc/nueva.html', context)
            else:
                # Manejo del archivo de imagen
                if 'image' in request.FILES:
                    form.instance.image = request.FILES['image']
                form.save()
                return redirect('pc')
        else:
            context = {
                'form': form
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_pc/nueva.html', context)




@login_required    
def detalles(request, id):
    if request.method == 'GET':
        pc = get_object_or_404(PC, id=id)
        mantenimientos = pc.mantenimientopc_set.all().order_by('-fecha', '-hora')
        form = PCForm(instance=pc)
        context = {
            'pc': pc,
            'form': form,
            'id': id,
            'mantenimientos': mantenimientos,
            }
        return render(request, 'SGE_pc/detalles.html', context)
    
    if request.method == 'POST':
        pc = get_object_or_404(PC, id=id)
        form = PCForm(instance=pc)
        form = PCForm(request.POST, request.FILES, instance=pc)

        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                mantenimientos = pc.mantenimientopc_set.all().order_by('-fecha', '-hora')
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'pc': pc,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            else:
                form.save()
                mantenimientos = pc.mantenimientopc_set.all().order_by('-fecha', '-hora')
                context = {
                    'pc': pc,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                return render(request, 'SGE_pc/detalles.html', context) 
        
        else:
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url)            




@login_required
def eliminar(request, id):
    pc = get_object_or_404(PC, id=id)
    pc.delete()
    return redirect('pc')

# fin de vistas generales---------------------------------------------    




@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoPC, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)




@login_required
def mantenimientos_pc_preventivos(request, id):
    if request.method == 'GET':
        pc = get_object_or_404(PC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=2) 
        mantenimientos = pc.mantenimientopc_set.filter(tipo=tipo_mantenimiento).order_by('-fecha', '-hora')
        context = {
            'pc': pc,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_pc/mantenimientos_preventivo.html', context) 




@login_required
def mod_mantenimiento_pc_preventivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoPC, id=id)
        pc = mantenimiento.pc
        form_mant = MantenimientoPCForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'pc': pc,
        }
        return render(request, 'SGE_pc/mod_mantenimiento_preventivo.html', context) 

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoPC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=2) 
        pc = mantenimiento.pc 
        form_mant = MantenimientoPCForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.pc = pc
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_pc_preventivos', id=pc.id)
        else:
            context = {
            'form_mant': form_mant,
            'pc': pc,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_pc/mod_mantenimiento_preventivo.html', context)                  

    return HttpResponse("Method Not Allowed", status=405) 




@login_required
def nuevo_mantenimiento_pc_preventivo(request, id):
    if request.method == 'GET':
        pc = get_object_or_404(PC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=2) 
        form_mant = MantenimientoPCForm()
        context = {
            'form_mant': form_mant,
            'pc': pc,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_pc/nuevo_mantenimiento_preventivo.html', context)
    
    if request.method == 'POST':
        pc = get_object_or_404(PC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=2) 
        form_mant = MantenimientoPCForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.pc = pc
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_pc_preventivos', id=pc.id)
        else:
            context = {
                'form_mant': form_mant,
                'pc': pc,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_pc/nuevo_mantenimiento_preventivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def mantenimientos_pc_correctivos(request, id):
    if request.method == 'GET':
        pc = get_object_or_404(PC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=1) 
        mantenimientos = pc.mantenimientopc_set.filter(tipo=tipo_mantenimiento).order_by('-fecha', '-hora')
        context = {
            'pc': pc,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_pc/mantenimientos_correctivo.html', context) 




@login_required
def mod_mantenimiento_pc_correctivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoPC, id=id)
        pc = mantenimiento.pc
        form_mant = MantenimientoPCForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'pc': pc,
        }
        return render(request, 'SGE_pc/mod_mantenimiento_correctivo.html', context) 

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoPC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=1) 
        pc = mantenimiento.pc 
        form_mant = MantenimientoPCForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.pc = pc
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_pc_correctivos', id=pc.id)
        else:
            context = {
            'form_mant': form_mant,
            'pc': pc,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_pc/mod_mantenimiento_correctivo.html', context)                  

    return HttpResponse("Method Not Allowed", status=405) 




@login_required
def nuevo_mantenimiento_pc_correctivo(request, id):
    if request.method == 'GET':
        pc = get_object_or_404(PC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=1) 
        form_mant = MantenimientoPCForm()
        context = {
            'form_mant': form_mant,
            'pc': pc,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_pc/nuevo_mantenimiento_correctivo.html', context)
    
    if request.method == 'POST':
        pc = get_object_or_404(PC, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=1) 
        form_mant = MantenimientoPCForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.pc = pc
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_pc_correctivos', id=pc.id)
        else:
            context = {
                'form_mant': form_mant,
                'pc': pc,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_pc/nuevo_mantenimiento_correctivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)    




@login_required
def documento_general_mantenimientos_pc(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoPC.objects.filter(fecha__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha__month=mes)

    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha', '-hora')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_equipos_de_cómputo_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_equipos_de_cómputo_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['E Cómputo', 'Tipo', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Operador', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.pc.nombre,
            mantenimiento.tipo.tipo,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha,
            mantenimiento.hora,
            mantenimiento.operador,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_preventivos_pc(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 2

    pc = get_object_or_404(PC, pk=id)
    mantenimientos = MantenimientoPC.objects.filter(pc=pc).order_by('-fecha', '-hora')

    if mes:
        mantenimientos = mantenimientos.filter(fecha__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}_{}.xlsx"'.format(pc.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}.xlsx"'.format(pc.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Operador', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha,
            mantenimiento.hora,
            mantenimiento.operador,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response    




@login_required
def documento_mantenimientos_correctivos_pc(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 1

    pc = get_object_or_404(PC, pk=id)
    mantenimientos = MantenimientoPC.objects.filter(pc=pc).order_by('-fecha', '-hora')

    if mes:
        mantenimientos = mantenimientos.filter(fecha__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivo_de_{}_{}_{}.xlsx"'.format(pc.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivo_de_{}_{}.xlsx"'.format(pc.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Operador', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha,
            mantenimiento.hora,
            mantenimiento.operador,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response   